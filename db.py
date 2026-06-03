# db.py
import sqlite3
import time
import json
import csv
import os
from config import DB_PATH, MIN_EVENT_INTERVAL

last_event_time = {}


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """
    Створює БД та таблицю подій.
    Якщо таблиця вже існує, перевіряє наявність потрібних колонок.
    """
    conn = get_connection()
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            frame INTEGER,
            object_class TEXT,
            event_type TEXT,
            bbox TEXT,
            confidence REAL,
            timestamp REAL
        )
    """)

    c.execute("PRAGMA table_info(events)")
    existing_columns = [row["name"] for row in c.fetchall()]

    # Міграції: додаємо відсутні колонки
    if "camera_id" not in existing_columns:
        c.execute("ALTER TABLE events ADD COLUMN camera_id TEXT DEFAULT 'CAM-01'")

    if "frame_path" not in existing_columns:
        c.execute("ALTER TABLE events ADD COLUMN frame_path TEXT DEFAULT NULL")

    for idx_sql in [
        "CREATE INDEX IF NOT EXISTS idx_events_timestamp ON events(timestamp)",
        "CREATE INDEX IF NOT EXISTS idx_events_object_class ON events(object_class)",
        "CREATE INDEX IF NOT EXISTS idx_events_camera_id ON events(camera_id)",
    ]:
        c.execute(idx_sql)

    conn.commit()
    conn.close()


def log_event(frame, object_class, event_type, bbox, confidence,
              camera_id="CAM-01", frame_path=None):
    """Логує подію в SQLite. bbox зберігається у JSON-форматі."""
    conn = get_connection()
    c = conn.cursor()

    timestamp = time.time()
    bbox_json = json.dumps(bbox, ensure_ascii=False)

    c.execute("""
        INSERT INTO events (
            frame, object_class, event_type, bbox, confidence,
            timestamp, camera_id, frame_path
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (frame, object_class, event_type, bbox_json, confidence,
          timestamp, camera_id, frame_path))

    conn.commit()
    conn.close()


def can_send_event(object_class, interval=None):
    """Антифлуд-фільтр."""
    threshold = interval if interval is not None else MIN_EVENT_INTERVAL
    now = time.time()
    last_time = last_event_time.get(object_class, 0)
    if now - last_time >= threshold:
        last_event_time[object_class] = now
        return True
    return False


def get_events(
    limit=300,
    object_class_filter=None,
    event_type_filter=None,
    min_confidence=0.0,
    date_from=None,
    date_to=None,
    camera_id_filter=None,
):
    """Повертає список подій з фільтрацією."""
    conn = get_connection()
    c = conn.cursor()

    query = "SELECT * FROM events WHERE 1=1"
    params = []

    if object_class_filter:
        query += " AND object_class = ?"
        params.append(object_class_filter)

    if event_type_filter:
        query += " AND event_type LIKE ?"
        params.append(f"%{event_type_filter}%")

    if min_confidence > 0.0:
        query += " AND confidence >= ?"
        params.append(min_confidence)

    if date_from is not None:
        query += " AND timestamp >= ?"
        params.append(date_from)

    if date_to is not None:
        query += " AND timestamp <= ?"
        params.append(date_to)

    if camera_id_filter:
        query += " AND camera_id = ?"
        params.append(camera_id_filter)

    query += " ORDER BY timestamp DESC LIMIT ?"
    params.append(limit)

    c.execute(query, params)
    rows = [dict(row) for row in c.fetchall()]
    conn.close()
    return rows


def get_stats():
    """Повертає статистику подій."""
    conn = get_connection()
    c = conn.cursor()
    now = time.time()

    c.execute("SELECT COUNT(*) FROM events")
    total = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM events WHERE timestamp >= ?", (now - 3600,))
    last_hour = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM events WHERE timestamp >= ?", (now - 86400,))
    last_day = c.fetchone()[0]

    c.execute("""
        SELECT object_class, COUNT(*) AS cnt
        FROM events GROUP BY object_class ORDER BY cnt DESC
    """)
    by_class = {row[0]: row[1] for row in c.fetchall()}

    # Активність по годинах (останні 24 год)
    c.execute("""
        SELECT strftime('%H', datetime(timestamp, 'unixepoch', 'localtime')) as hour,
               COUNT(*) as cnt
        FROM events
        WHERE timestamp >= ?
        GROUP BY hour ORDER BY hour
    """, (now - 86400,))
    by_hour = {row[0]: row[1] for row in c.fetchall()}

    conn.close()
    return {
        "total": total,
        "last_hour": last_hour,
        "last_day": last_day,
        "by_class": by_class,
        "by_hour": by_hour,
    }


def delete_event(event_id):
    """Видаляє одну подію за ID."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM events WHERE id = ?", (event_id,))
    conn.commit()
    conn.close()


def clear_all_events():
    """Очищає всі події з БД."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM events")
    conn.commit()
    conn.close()


def export_to_csv(path: str, **filters) -> int:
    """Експортує події у CSV. Повертає кількість рядків."""
    import datetime as dt
    events = get_events(limit=10000, **filters)
    if not events:
        return 0

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "timestamp_str", "object_class", "event_type",
            "confidence", "frame", "camera_id", "frame_path"
        ])
        writer.writeheader()
        for ev in events:
            ts = ev.get("timestamp", 0)
            try:
                ts_str = dt.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                ts_str = ""
            writer.writerow({
                "id": ev.get("id"),
                "timestamp_str": ts_str,
                "object_class": ev.get("object_class"),
                "event_type": ev.get("event_type"),
                "confidence": f"{float(ev.get('confidence', 0)):.2%}",
                "frame": ev.get("frame"),
                "camera_id": ev.get("camera_id"),
                "frame_path": ev.get("frame_path") or "",
            })
    return len(events)


def export_to_excel(path: str, **filters) -> int:
    """Експортує події у Excel (.xlsx). Повертає кількість рядків."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("Встановіть openpyxl: pip install openpyxl")

    import datetime as dt
    events = get_events(limit=10000, **filters)
    if not events:
        return 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Події"

    headers = ["ID", "Час", "Клас об'єкта", "Тип події", "Впевненість", "Кадр", "Камера", "Кадр файл"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, ev in enumerate(events, 2):
        ts = ev.get("timestamp", 0)
        try:
            ts_str = dt.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            ts_str = ""
        conf = float(ev.get("confidence", 0))
        ws.append([
            ev.get("id"), ts_str, ev.get("object_class"),
            ev.get("event_type"), f"{conf:.0%}",
            ev.get("frame"), ev.get("camera_id"),
            ev.get("frame_path") or "",
        ])
        # Колір рядка по confidence
        if conf >= 0.8:
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif conf >= 0.5:
            fill = PatternFill("solid", fgColor="FFEB9C")
        else:
            fill = PatternFill("solid", fgColor="FFC7CE")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = fill

    # Авторозмір колонок
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(path)
    return len(events)